"""
backfill_fwa_from_sheet.py

Reads FWA values from Franchise_baseball_metrics.xlsx (weeks 1-11)
and upserts into pipeline.weekly_metric_snapshots.

Overwrites any existing FWA for these weeks.
FER is intentionally ignored — will be computed separately.
Week 12 is skipped (partial/broken — only More Defiant Jazz filled in).

Upsert key: (player_id, week_number, season_year)

Run from repo root:
    python backfill_fwa_from_sheet.py --xlsx /path/to/Franchise_baseball_metrics.xlsx
"""

import os
import sys
import argparse
import uuid
import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUPABASE_URL = os.environ["SUPABASE_URL"]          # e.g. https://seqvzektwxxypdcqgtve.supabase.co
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]  # service role key
SEASON_YEAR  = 2026
WEEKS        = list(range(1, 12))   # 1-11 inclusive; 12 skipped (partial)

HEADERS = {
    "apikey":          SUPABASE_KEY,
    "Authorization":   f"Bearer {SUPABASE_KEY}",
    "Content-Type":    "application/json",
    "Accept-Profile":  "pipeline",
    "Content-Profile": "pipeline",
    "Prefer":          "resolution=merge-duplicates",
}

# ---------------------------------------------------------------------------
# Team name mapping: spreadsheet casing → DB team_name (exact match)
# ---------------------------------------------------------------------------
TEAM_NAME_MAP = {
    "More Defiant Jazz":       "More Defiant Jazz",
    "All Betts Are Off":       "All Betts are Off",
    "Ass Cannons":             "Ass Cannons",
    "Boston Stink Sox":        "Boston Stink Sox",
    "Down By The Schoolyard":  "Down by the Schoolyard",
    "Greene Brown and Schlitty": "Greene Brown and Schlitty",
    "I am the Breg-Man":       "I am the Breg-man",
    "Kekambas":                "Kekambas",
    "My Roman Empire":         "My Roman Empire",
    "Ronald's PlayPlace":      "Ronald's PlayPlace",
    "Sho-Time":                "Sho-Time",
    "honey nuts":              "Jackson County OrangTurangs",
    "Honey Nuts":              "Jackson County OrangTurangs",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sb_get(path, params=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    r = requests.get(url, headers=HEADERS, params=params)
    r.raise_for_status()
    return r.json()


def sb_upsert(path, rows, batch_size=200):
    """POST with prefer merge-duplicates (upsert). Returns total rows written."""
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    written = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i+batch_size]
        r = requests.post(url, headers=HEADERS, json=batch)
        if not r.ok:
            print(f"  ERROR batch {i//batch_size}: {r.status_code} {r.text[:300]}")
            r.raise_for_status()
        written += len(batch)
    return written


def load_teams(season_year):
    """
    Returns two dicts:
      db_name_to_id  : "More Defiant Jazz" -> uuid
      sheet_name_to_id: "More Defiant Jazz" -> uuid  (via TEAM_NAME_MAP)
    """
    # Pull team_seasons joined to teams for this season
    rows = sb_get(
        "team_seasons",
        params={
            "select": "team_id,teams(team_name),seasons(year)",
            "seasons.year": f"eq.{season_year}",
        }
    )

    # Simpler: pull all teams then filter by season via separate query
    # Actually pull teams via teams table directly, filtering by season
    # Use execute SQL pattern via REST
    # Simpler approach: pull all teams and all team_seasons, join in Python
    all_teams = sb_get("teams", params={"select": "id,team_name,yahoo_team_id"})
    seasons   = sb_get("seasons", params={"select": "id,year", "year": f"eq.{season_year}"})

    if not seasons:
        raise ValueError(f"No season found for year {season_year}")
    season_id = seasons[0]["id"]

    team_seasons = sb_get(
        "team_seasons",
        params={"select": "team_id", "season_id": f"eq.{season_id}"}
    )
    active_team_ids = {ts["team_id"] for ts in team_seasons}

    db_name_to_id = {
        t["team_name"]: t["id"]
        for t in all_teams
        if t["id"] in active_team_ids
    }

    # Build sheet-name → uuid map
    sheet_name_to_id = {}
    for sheet_name, db_name in TEAM_NAME_MAP.items():
        if db_name in db_name_to_id:
            sheet_name_to_id[sheet_name] = db_name_to_id[db_name]
        else:
            print(f"  WARNING: DB team not found for sheet name '{sheet_name}' -> '{db_name}'")

    return sheet_name_to_id


def load_players():
    """Returns dict: normalized_full_name -> player_id uuid"""
    all_players = sb_get(
        "players",
        params={"select": "id,first_name,last_name", "limit": 10000}
    )
    name_to_id = {}
    for p in all_players:
        full = f"{p['first_name']} {p['last_name']}".strip()
        norm = normalize_name(full)
        name_to_id[norm] = p["id"]
    return name_to_id


def normalize_name(name):
    """Lowercase, strip accents/special chars for fuzzy matching."""
    import unicodedata
    name = unicodedata.normalize("NFD", name)
    name = "".join(c for c in name if unicodedata.category(c) != "Mn")
    name = name.lower().strip()
    # Remove suffixes
    for suffix in [" jr.", " jr", " sr.", " sr", " iii", " ii"]:
        if name.endswith(suffix):
            name = name[:-len(suffix)].strip()
    return name


def extract_week(wb, week_num):
    """
    Returns list of dicts: {player_name, team_sheet_name, fwa}
    Only rows with a non-None, non-zero FWA and a valid player name.
    """
    sheet_name = f"Week {week_num}"
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row_idx = None
    for i, row in enumerate(all_rows):
        row_strs = [str(c) if c is not None else "" for c in row]
        if "Player" in row_strs and "FWA" in row_strs:
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  Week {week_num}: could not find header row, skipping")
        return []

    header = all_rows[header_row_idx]
    col_player = col_team = col_fwa = None
    for i, h in enumerate(header):
        if h == "Player":  col_player = i
        if h == "Team":    col_team = i
        if h == "FWA":     col_fwa = i

    if col_player is None or col_fwa is None:
        print(f"  Week {week_num}: missing Player or FWA column, skipping")
        return []

    if col_team is None:
        print(f"  Week {week_num}: no Team column found")

    rows_out = []
    for row in all_rows[header_row_idx + 1:]:
        player = row[col_player] if col_player < len(row) else None
        team   = row[col_team]   if col_team is not None and col_team < len(row) else None
        fwa    = row[col_fwa]    if col_fwa < len(row) else None

        if not (player and isinstance(player, str) and player.strip()):
            continue
        if not isinstance(fwa, (int, float)):
            continue
        # Include zero-FWA rows — a player can legitimately have 0 FWA
        # (e.g. benched all week, no contribution)
        rows_out.append({
            "player_name":      player.strip(),
            "team_sheet_name":  team.strip() if isinstance(team, str) else None,
            "fwa":              float(fwa),
        })

    return rows_out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to Franchise_baseball_metrics.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Print counts without writing")
    args = parser.parse_args()

    print("Loading workbook...")
    wb = load_workbook(args.xlsx, read_only=True, data_only=True)

    print("Loading teams from Supabase...")
    sheet_name_to_team_id = load_teams(SEASON_YEAR)
    print(f"  {len(sheet_name_to_team_id)} teams mapped")

    print("Loading players from Supabase...")
    norm_name_to_player_id = load_players()
    print(f"  {len(norm_name_to_player_id)} players in DB")

    grand_total = 0
    unmatched_players = {}   # name -> list of weeks
    unmatched_teams   = {}   # team -> list of weeks

    for week_num in WEEKS:
        print(f"\n--- Week {week_num} ---")
        sheet_rows = extract_week(wb, week_num)
        print(f"  {len(sheet_rows)} player rows in sheet")

        upsert_rows = []
        skipped_no_player = 0
        skipped_no_team   = 0

        for sr in sheet_rows:
            # Resolve player
            norm = normalize_name(sr["player_name"])
            player_id = norm_name_to_player_id.get(norm)
            if player_id is None:
                skipped_no_player += 1
                if sr["player_name"] not in unmatched_players:
                    unmatched_players[sr["player_name"]] = []
                unmatched_players[sr["player_name"]].append(week_num)
                continue

            # Resolve team
            team_id = None
            if sr["team_sheet_name"]:
                team_id = sheet_name_to_team_id.get(sr["team_sheet_name"])
                if team_id is None:
                    skipped_no_team += 1
                    if sr["team_sheet_name"] not in unmatched_teams:
                        unmatched_teams[sr["team_sheet_name"]] = []
                    unmatched_teams[sr["team_sheet_name"]].append(week_num)
                    continue

            upsert_rows.append({
                "id":          str(uuid.uuid4()),
                "player_id":   player_id,
                "team_id":     team_id,
                "week_number": week_num,
                "season_year": SEASON_YEAR,
                "fwa":         round(sr["fwa"], 6),
                "fer":         None,   # FER intentionally excluded
                "fer_grade":   None,
                "stat_basis":  "manual_sheet",
                "is_locked":   True,
            })

        print(f"  {len(upsert_rows)} rows ready to upsert")
        print(f"  {skipped_no_player} skipped (player not in DB)")
        print(f"  {skipped_no_team} skipped (team not in DB)")

        if not args.dry_run and upsert_rows:
            written = sb_upsert("weekly_metric_snapshots", upsert_rows)
            print(f"  ✓ {written} rows written")
            grand_total += written
        elif args.dry_run:
            # Verify FWA checksum for this week
            total_fwa = sum(r["fwa"] for r in upsert_rows)
            print(f"  [DRY RUN] total FWA this week: {total_fwa:.4f}")

    print(f"\n{'='*50}")
    if not args.dry_run:
        print(f"DONE — {grand_total} total rows written across weeks {WEEKS[0]}-{WEEKS[-1]}")
    else:
        print("DRY RUN COMPLETE — nothing written")

    if unmatched_players:
        print(f"\nUnmatched players ({len(unmatched_players)} unique):")
        for name, weeks in sorted(unmatched_players.items()):
            print(f"  '{name}' — weeks {weeks}")

    if unmatched_teams:
        print(f"\nUnmatched teams ({len(unmatched_teams)} unique):")
        for name, weeks in sorted(unmatched_teams.items()):
            print(f"  '{name}' — weeks {weeks}")


if __name__ == "__main__":
    main()
